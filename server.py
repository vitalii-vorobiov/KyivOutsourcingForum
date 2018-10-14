from flask import Flask, render_template, request, redirect
from openpyxl import load_workbook
import os
import ast

app = Flask(__name__)

APP_ROOT = os.path.dirname(os.path.abspath(__file__))   # refers to application_top
APP_STATIC = os.path.join(APP_ROOT, 'static')

@app.route('/')
def hello():
    wb = load_workbook(filename=os.path.join(APP_STATIC, 'reservations.xlsx'))
    sheet_ranges = wb['Sheet1']

    lst = []

    for row in sheet_ranges.rows:
        lst.append([row[0].value, [row[i].value for i in range(1, len(row))]])

    final_lst = []

    for i in range(1, len(lst)):
        final_lst.append([])

        final_lst[i-1].append(lst[i][0])
        final_lst[i-1].append([])

        for j in range(len(lst[i][1])):
            if lst[i][1][j] == "free":
                final_lst[i-1][1].append(lst[0][1][j])
            elif lst[i][1][j] != "free" and lst[i][1][j] != None:
                final_lst[i - 1][1].append("Reserved")
    wb.close()
    return render_template("index.html", lst=final_lst)


@app.route('/postmethod', methods=['POST'])
def post_javascript_data():
    res_data = request.form['canvas_data']
    res_data = ast.literal_eval(res_data)


    wb = load_workbook(filename=os.path.join(APP_STATIC, 'reservations.xlsx'))
    sheet_ranges = wb['Sheet1']

    lst = []

    for row in sheet_ranges.rows:
        if row[0].value == res_data[0]:
            for cell in row:
                if sheet_ranges[f"{cell.column}1"].value == res_data[1] and sheet_ranges[f"{cell.column}{cell.row}"].value == "free":
                    with open(os.path.join(APP_STATIC, 'users.txt')) as f:
                        for line in f.readlines():
                            if res_data[2] in line:
                                cell.value = res_data[2]

    wb.save(os.path.join(APP_STATIC, 'reservations.xlsx'))
    wb.close()
    wb = load_workbook(filename=os.path.join(APP_STATIC, 'reservations.xlsx'))
    sheet_ranges = wb['Sheet1']

    for row in sheet_ranges.rows:
        lst.append([row[0].value, [row[i].value for i in range(1, len(row))]])

    final_lst = []

    for i in range(1, len(lst)):
        final_lst.append([])

        final_lst[i - 1].append(lst[i][0])
        final_lst[i - 1].append([])

        for j in range(len(lst[i][1])):
            if lst[i][1][j] == "free":
                final_lst[i - 1][1].append(lst[0][1][j])
            elif lst[i][1][j] != "free" and lst[i][1][j] != None:
                final_lst[i - 1][1].append("Reserved")

    wb.close()
    return redirect(hello)

if __name__ == '__main__':
    app.run(debug=True, use_reloader=True)